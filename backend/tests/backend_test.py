"""Backend regression tests for مخيم العائدين (Refugee Camp app).

Covers: auth, role enforcement, family-fields, families, aid-types, aid-records,
stats, Excel export/template/import.
"""
import io
import os
import time
import pytest
import requests
import openpyxl

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
# Backend URL is exposed via the frontend env in this preview environment
if not BASE_URL:
    BASE_URL = "https://aid-database.preview.emergentagent.com"

ADMIN_EMAIL = "admin@camp.com"
ADMIN_PASSWORD = "admin123"

TEST_PREFIX = f"TEST_{int(time.time())}_"

# ----------------------- Fixtures -----------------------
@pytest.fixture(scope="session")
def admin_token():
    r = requests.post(f"{BASE_URL}/api/auth/login",
                      json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}, timeout=20)
    assert r.status_code == 200, f"admin login failed: {r.status_code} {r.text}"
    return r.json()["token"]


@pytest.fixture(scope="session")
def admin_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}", "Content-Type": "application/json"}


@pytest.fixture(scope="session")
def staff_credentials(admin_headers):
    """Create a staff user once, return its login credentials."""
    email = f"{TEST_PREFIX}staff@test.com".lower()
    password = "staffpass123"
    body = {"email": email, "password": password, "name": "TEST staff", "role": "staff"}
    r = requests.post(f"{BASE_URL}/api/users", json=body, headers=admin_headers, timeout=20)
    assert r.status_code == 200, f"staff create failed: {r.status_code} {r.text}"
    user_id = r.json()["id"]
    yield {"email": email, "password": password, "id": user_id}
    requests.delete(f"{BASE_URL}/api/users/{user_id}", headers=admin_headers, timeout=20)


@pytest.fixture(scope="session")
def staff_token(staff_credentials):
    r = requests.post(f"{BASE_URL}/api/auth/login",
                      json={"email": staff_credentials["email"],
                            "password": staff_credentials["password"]}, timeout=20)
    assert r.status_code == 200
    return r.json()["token"]


@pytest.fixture(scope="session")
def staff_headers(staff_token):
    return {"Authorization": f"Bearer {staff_token}", "Content-Type": "application/json"}


# ----------------------- Auth -----------------------
class TestAuth:
    def test_login_success(self):
        r = requests.post(f"{BASE_URL}/api/auth/login",
                          json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}, timeout=20)
        assert r.status_code == 200
        data = r.json()
        assert "token" in data and isinstance(data["token"], str) and len(data["token"]) > 10
        assert data["user"]["email"] == ADMIN_EMAIL
        assert data["user"]["role"] == "admin"

    def test_login_invalid_password(self):
        r = requests.post(f"{BASE_URL}/api/auth/login",
                          json={"email": ADMIN_EMAIL, "password": "wrongpass"}, timeout=20)
        assert r.status_code == 401
        assert "detail" in r.json()

    def test_login_unknown_email(self):
        r = requests.post(f"{BASE_URL}/api/auth/login",
                          json={"email": "nobody@nope.com", "password": "x"}, timeout=20)
        assert r.status_code == 401

    def test_me_with_token(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/auth/me", headers=admin_headers, timeout=20)
        assert r.status_code == 200
        assert r.json()["email"] == ADMIN_EMAIL
        assert r.json()["role"] == "admin"

    def test_me_without_token(self):
        r = requests.get(f"{BASE_URL}/api/auth/me", timeout=20)
        assert r.status_code == 401


# ----------------------- Role enforcement -----------------------
class TestRoleEnforcement:
    """Staff token must receive 403 on admin-only endpoints."""

    def test_staff_cannot_create_field(self, staff_headers):
        r = requests.post(f"{BASE_URL}/api/family-fields",
                          json={"label": "x", "type": "text", "order": 0},
                          headers=staff_headers, timeout=20)
        assert r.status_code == 403

    def test_staff_cannot_create_aid_type(self, staff_headers):
        r = requests.post(f"{BASE_URL}/api/aid-types",
                          json={"name": "x"}, headers=staff_headers, timeout=20)
        assert r.status_code == 403

    def test_staff_cannot_create_user(self, staff_headers):
        r = requests.post(f"{BASE_URL}/api/users",
                          json={"email": "a@b.c", "password": "p", "name": "n", "role": "staff"},
                          headers=staff_headers, timeout=20)
        assert r.status_code == 403

    def test_staff_cannot_delete_family(self, staff_headers, admin_headers):
        # admin creates a family
        r = requests.post(f"{BASE_URL}/api/families",
                          json={"data": {"name": "TEST role"}},
                          headers=admin_headers, timeout=20)
        assert r.status_code == 200
        fid = r.json()["id"]
        # staff attempts to delete
        r2 = requests.delete(f"{BASE_URL}/api/families/{fid}", headers=staff_headers, timeout=20)
        assert r2.status_code == 403
        # cleanup
        requests.delete(f"{BASE_URL}/api/families/{fid}", headers=admin_headers, timeout=20)


# ----------------------- Family fields -----------------------
class TestFamilyFields:
    created_ids = []

    def test_create_field_persists(self, admin_headers):
        label = f"{TEST_PREFIX}اسم رب الأسرة"
        r = requests.post(f"{BASE_URL}/api/family-fields",
                          json={"label": label, "type": "text", "order": 1},
                          headers=admin_headers, timeout=20)
        assert r.status_code == 200
        rec = r.json()
        assert rec["label"] == label
        assert rec["type"] == "text"
        assert "id" in rec and "key" in rec
        TestFamilyFields.created_ids.append(rec["id"])
        # GET verify
        r2 = requests.get(f"{BASE_URL}/api/family-fields", headers=admin_headers, timeout=20)
        assert r2.status_code == 200
        assert any(f["id"] == rec["id"] for f in r2.json())

    def test_delete_field(self, admin_headers):
        # create a throwaway field
        r = requests.post(f"{BASE_URL}/api/family-fields",
                          json={"label": f"{TEST_PREFIX}del", "type": "text", "order": 99},
                          headers=admin_headers, timeout=20)
        fid = r.json()["id"]
        d = requests.delete(f"{BASE_URL}/api/family-fields/{fid}", headers=admin_headers, timeout=20)
        assert d.status_code == 200
        r2 = requests.get(f"{BASE_URL}/api/family-fields", headers=admin_headers, timeout=20)
        assert all(f["id"] != fid for f in r2.json())


# ----------------------- Aid Types -----------------------
class TestAidTypes:
    def test_create_and_list(self, admin_headers):
        name = f"{TEST_PREFIX}سلة غذائية"
        r = requests.post(f"{BASE_URL}/api/aid-types",
                          json={"name": name, "description": "test"},
                          headers=admin_headers, timeout=20)
        assert r.status_code == 200
        rid = r.json()["id"]
        assert r.json()["name"] == name

        r2 = requests.get(f"{BASE_URL}/api/aid-types", headers=admin_headers, timeout=20)
        assert r2.status_code == 200
        assert any(a["id"] == rid for a in r2.json())

        d = requests.delete(f"{BASE_URL}/api/aid-types/{rid}", headers=admin_headers, timeout=20)
        assert d.status_code == 200


# ----------------------- Families + Aid Records full flow -----------------------
class TestFamiliesAidFlow:
    def test_full_flow(self, admin_headers):
        # ensure a field exists
        flabel = f"{TEST_PREFIX}الاسم"
        fr = requests.post(f"{BASE_URL}/api/family-fields",
                           json={"label": flabel, "type": "text", "order": 0},
                           headers=admin_headers, timeout=20)
        assert fr.status_code == 200
        field = fr.json()
        fkey = field["key"]

        # create aid type
        atr = requests.post(f"{BASE_URL}/api/aid-types",
                            json={"name": f"{TEST_PREFIX}طرد"},
                            headers=admin_headers, timeout=20)
        assert atr.status_code == 200
        aid_type_id = atr.json()["id"]

        # create family using that field
        fam_name = f"{TEST_PREFIX}عائلة الاختبار"
        cr = requests.post(f"{BASE_URL}/api/families",
                           json={"data": {fkey: fam_name}},
                           headers=admin_headers, timeout=20)
        assert cr.status_code == 200
        family_id = cr.json()["id"]
        assert cr.json()["data"][fkey] == fam_name

        # GET families - verify presence
        lr = requests.get(f"{BASE_URL}/api/families", headers=admin_headers, timeout=20)
        assert lr.status_code == 200
        assert any(f["id"] == family_id for f in lr.json())

        # create aid record
        ar = requests.post(f"{BASE_URL}/api/aid-records",
                           json={"family_id": family_id, "aid_type_id": aid_type_id,
                                 "date": "2026-01-15", "quantity": "2", "notes": "TEST"},
                           headers=admin_headers, timeout=20)
        assert ar.status_code == 200
        record = ar.json()
        assert record["family_id"] == family_id
        assert record["aid_type_name"]  # denormalized

        # list aid records filtered by family
        lar = requests.get(f"{BASE_URL}/api/aid-records?family_id={family_id}",
                           headers=admin_headers, timeout=20)
        assert lar.status_code == 200
        assert any(r["id"] == record["id"] for r in lar.json())

        # stats includes our family
        sr = requests.get(f"{BASE_URL}/api/stats", headers=admin_headers, timeout=20)
        assert sr.status_code == 200
        s = sr.json()
        assert s["total_families"] >= 1 and s["total_aid_records"] >= 1
        assert isinstance(s["aid_by_type"], list)

        # cleanup
        requests.delete(f"{BASE_URL}/api/aid-records/{record['id']}", headers=admin_headers, timeout=20)
        requests.delete(f"{BASE_URL}/api/families/{family_id}", headers=admin_headers, timeout=20)
        requests.delete(f"{BASE_URL}/api/aid-types/{aid_type_id}", headers=admin_headers, timeout=20)
        requests.delete(f"{BASE_URL}/api/family-fields/{field['id']}", headers=admin_headers, timeout=20)


# ----------------------- Excel export / template / import -----------------------
class TestExcel:
    def test_export_families(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/families/export", headers=admin_headers, timeout=30)
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct or "xlsx" in ct, f"unexpected content-type: {ct}"
        # verify it is a valid xlsx
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert wb.active is not None

    def test_template_families(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/families/template", headers=admin_headers, timeout=30)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")

    def test_export_aid_records(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/aid-records/export", headers=admin_headers, timeout=30)
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # header row in Arabic
        headers_row = [c.value for c in next(ws.iter_rows(max_row=1))]
        assert "العائلة" in headers_row

    def test_import_families_round_trip(self, admin_headers):
        # create a field
        label = f"{TEST_PREFIX}اسم_استيراد"
        fr = requests.post(f"{BASE_URL}/api/family-fields",
                           json={"label": label, "type": "text", "order": 50},
                           headers=admin_headers, timeout=20)
        assert fr.status_code == 200
        field_id = fr.json()["id"]
        fkey = fr.json()["key"]

        # build xlsx with label header
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([label])
        ws.append([f"{TEST_PREFIX}عائلة_1"])
        ws.append([f"{TEST_PREFIX}عائلة_2"])
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)

        token = admin_headers["Authorization"]
        files = {"file": ("import.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/families/import",
                          files=files, headers={"Authorization": token}, timeout=30)
        assert r.status_code == 200, f"import failed: {r.status_code} {r.text}"
        assert r.json()["imported"] >= 2

        # verify families exist
        lr = requests.get(f"{BASE_URL}/api/families", headers=admin_headers, timeout=20)
        created = [f for f in lr.json()
                   if isinstance(f.get("data"), dict) and
                   str(f["data"].get(fkey, "")).startswith(f"{TEST_PREFIX}عائلة")]
        assert len(created) >= 2
        # cleanup
        for f in created:
            requests.delete(f"{BASE_URL}/api/families/{f['id']}", headers=admin_headers, timeout=20)
        requests.delete(f"{BASE_URL}/api/family-fields/{field_id}", headers=admin_headers, timeout=20)


# ----------------------- Users CRUD -----------------------
class TestUsers:
    def test_list_users_admin(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/users", headers=admin_headers, timeout=20)
        assert r.status_code == 200
        assert any(u["email"] == ADMIN_EMAIL for u in r.json())

    def test_staff_login_and_me(self, staff_credentials):
        r = requests.post(f"{BASE_URL}/api/auth/login",
                          json={"email": staff_credentials["email"],
                                "password": staff_credentials["password"]}, timeout=20)
        assert r.status_code == 200
        assert r.json()["user"]["role"] == "staff"
