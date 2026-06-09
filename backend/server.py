from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import io
import re
import json
import logging
from datetime import datetime, timezone
from typing import List, Optional, Any, Dict

from fastapi import FastAPI, APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
import openpyxl

import firebase_service as fs
import auth

# Initialize Firebase
fs.init_firebase()

app = FastAPI(title="مخيم العائدين API")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("camp")


# ─── Arabic fuzzy-matching helpers ──────────────────────────────────────────

def normalize_arabic(text: str) -> str:
    """Normalize Arabic text: remove diacritics + normalize letter variants."""
    if not text:
        return ""
    text = str(text).strip()
    # Remove Arabic diacritics (tashkeel / harakaat)
    text = re.sub(r'[\u0610-\u061A\u064B-\u065F\u0670\u06D6-\u06DC\u06DF-\u06E4\u06E7\u06E8\u06EA-\u06ED]', '', text)
    # Normalize alef variants → ا
    text = re.sub(r'[أإآٱ]', 'ا', text)
    # Normalize ta marbuta → ه
    text = re.sub(r'ة', 'ه', text)
    # Normalize alef maqsura → ي
    text = re.sub(r'ى', 'ي', text)
    # Collapse spaces
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def match_score(val_a: str, val_b: str) -> float:
    """Return 0.0-1.0 match score between two Arabic strings."""
    a = normalize_arabic(val_a)
    b = normalize_arabic(val_b)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    # Substring match (one name is part of the other)
    if a in b or b in a:
        return 0.85
    # Word-level matching
    wa = set(a.split())
    wb = set(b.split())
    common = len(wa & wb)
    if common == 0:
        return 0.0
    total = max(len(wa), len(wb))
    score = common / total
    # Boost if shorter name is a full subset of the longer
    shorter, longer = (wa, wb) if len(wa) <= len(wb) else (wb, wa)
    if shorter.issubset(longer):
        score = min(1.0, score + 0.15)
    return score


def find_best_family(families: list, field_key: str, query: str, threshold: float = 0.60):
    """Find the best-matching family for a query value using fuzzy matching."""
    best_score = 0.0
    best_fam = None
    q_norm = normalize_arabic(query)
    for fam in families:
        fam_val = str(fam.get("data", {}).get(field_key, "")).strip()
        if not fam_val:
            continue
        score = match_score(q_norm, fam_val)
        if score > best_score:
            best_score = score
            best_fam = fam
    if best_score >= threshold:
        return best_fam, best_score
    return None, best_score


def now_iso():
    return datetime.now(timezone.utc).isoformat()


# ----------------------- Models -----------------------
class LoginInput(BaseModel):
    email: str
    password: str


class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: str = "staff"


class RoleUpdate(BaseModel):
    role: str


class FieldCreate(BaseModel):
    label: str
    key: Optional[str] = None
    type: str = "text"
    order: int = 0


class FamilyInput(BaseModel):
    data: Dict[str, Any]


class AidTypeCreate(BaseModel):
    name: str
    description: Optional[str] = ""


class AidRecordInput(BaseModel):
    family_id: str
    aid_type_id: str
    date: str
    quantity: Optional[str] = ""
    notes: Optional[str] = ""



class IndividualMemberCreate(BaseModel):
    family_id: str            # references /families collection
    name: str
    id_number: str
    birth_date: str
    relation: str
    gender: str
    notes: Optional[str] = ""


# ----------------------- Individual Members (detailed) -----------------------
@api.get("/individual-members/count")
def count_individual_members(family_id: str, user: dict = Depends(auth.get_current_user)):
    """Return how many individual members are registered for a given family_id."""
    members = fs.list_records("individual_members")
    count = sum(1 for m in members if m.get("family_id") == family_id)
    return {"count": count}


@api.get("/individual-members")
def get_individual_members(family_id: Optional[str] = None, user: dict = Depends(auth.get_current_user)):
    members = fs.list_records("individual_members")
    if family_id:
        members = [m for m in members if m.get("family_id") == family_id]
    return sorted(members, key=lambda m: m.get("created_at", ""), reverse=True)


@api.post("/individual-members")
def add_individual_member(body: IndividualMemberCreate, user: dict = Depends(auth.get_current_user)):
    rec = fs.push("individual_members", {
        **body.model_dump(),
        "created_at": now_iso(),
        "created_by": user.get("name"),
    })
    return rec


@api.put("/individual-members/{member_id}")
def update_individual_member(member_id: str, body: IndividualMemberCreate, user: dict = Depends(auth.get_current_user)):
    if not fs.get_record("individual_members", member_id):
        raise HTTPException(status_code=404, detail="الفرد غير موجود")
    fs.update_record("individual_members", member_id, {**body.model_dump(), "updated_at": now_iso()})
    return fs.get_record("individual_members", member_id)


@api.delete("/individual-members/all")
def delete_all_individual_members(admin: dict = Depends(auth.require_admin)):
    members = fs.list_records("individual_members")
    count = len(members)
    fs.ref("individual_members").delete()
    return {"ok": True, "deleted": count}


@api.delete("/individual-members/{member_id}")
def delete_individual_member(member_id: str, user: dict = Depends(auth.get_current_user)):
    fs.delete_record("individual_members", member_id)
    return {"ok": True}


# ----------------------- Auth -----------------------
@api.post("/auth/login")
def login(body: LoginInput):
    user = auth.find_user_by_email(body.email)
    if not user or not auth.verify_password(body.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="البريد الإلكتروني أو كلمة المرور غير صحيحة")
    token = auth.create_access_token(user["id"], user["email"], user.get("role", "staff"))
    return {
        "token": token,
        "user": {"id": user["id"], "email": user["email"], "name": user.get("name"), "role": user.get("role")},
    }


@api.get("/auth/me")
def me(user: dict = Depends(auth.get_current_user)):
    return {"id": user["id"], "email": user["email"], "name": user.get("name"), "role": user.get("role")}


class FirebaseTokenInput(BaseModel):
    id_token: str


@api.post("/auth/firebase-login")
def firebase_login(body: FirebaseTokenInput):
    """Verify a Firebase ID token (obtained from Firebase Auth in the frontend)."""
    try:
        import firebase_admin.auth as fb_auth
        decoded = fb_auth.verify_id_token(body.id_token)
        email = decoded.get("email", "").lower()
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"رمز Firebase غير صالح: {str(e)}")

    # Look up user in Realtime DB by email
    user = auth.find_user_by_email(email)
    if not user:
        # Auto-create as staff if not found
        user_data = fs.push("users", {
            "email": email,
            "password_hash": "",
            "name": decoded.get("name") or email.split("@")[0],
            "role": "admin",
            "created_at": now_iso(),
        })
        user = user_data

    token = auth.create_access_token(user["id"], user["email"], user.get("role", "staff"))
    return {
        "token": token,
        "user": {"id": user["id"], "email": user["email"], "name": user.get("name"), "role": user.get("role")},
    }


# ----------------------- Users (admin) -----------------------
@api.get("/users")
def list_users(admin: dict = Depends(auth.require_admin)):
    users = fs.list_records("users")
    return [{"id": u["id"], "email": u["email"], "name": u.get("name"), "role": u.get("role"),
             "created_at": u.get("created_at")} for u in users]


@api.post("/users")
def create_user(body: UserCreate, admin: dict = Depends(auth.require_admin)):
    if body.role not in ("admin", "staff"):
        raise HTTPException(status_code=400, detail="صلاحية غير صالحة")
    if auth.find_user_by_email(body.email):
        raise HTTPException(status_code=400, detail="البريد الإلكتروني مستخدم بالفعل")
    rec = fs.push("users", {
        "email": body.email.strip().lower(),
        "password_hash": auth.hash_password(body.password),
        "name": body.name,
        "role": body.role,
        "created_at": now_iso(),
    })
    return {"id": rec["id"], "email": rec["email"], "name": rec["name"], "role": rec["role"]}


@api.delete("/users/{user_id}")
def delete_user(user_id: str, admin: dict = Depends(auth.require_admin)):
    if user_id == admin["id"]:
        raise HTTPException(status_code=400, detail="لا يمكنك حذف حسابك الخاص")
    fs.delete_record("users", user_id)
    return {"ok": True}


@api.put("/users/{user_id}/role")
def update_user_role(user_id: str, body: RoleUpdate, admin: dict = Depends(auth.require_admin)):
    if body.role not in ("admin", "staff"):
        raise HTTPException(status_code=400, detail="صلاحية غير صالحة")
    if user_id == admin["id"]:
        raise HTTPException(status_code=400, detail="لا يمكنك تغيير صلاحية حسابك الخاص")
    target = fs.get_record("users", user_id)
    if not target:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    fs.update_record("users", user_id, {"role": body.role})
    updated = fs.get_record("users", user_id)
    return {"id": updated["id"], "email": updated["email"], "name": updated.get("name"), "role": updated.get("role")}


# ----------------------- Family Fields -----------------------
@api.get("/family-fields")
def get_fields(user: dict = Depends(auth.get_current_user)):
    fields = fs.list_records("family_fields")
    return sorted(fields, key=lambda f: f.get("order", 0))


@api.post("/family-fields")
def add_field(body: FieldCreate, admin: dict = Depends(auth.require_admin)):
    key = body.key or f"f_{int(datetime.now().timestamp()*1000)}"
    rec = fs.push("family_fields", {
        "label": body.label,
        "key": key,
        "type": body.type,
        "order": body.order,
        "created_at": now_iso(),
    })
    return rec


@api.delete("/family-fields/{field_id}")
def delete_field(field_id: str, admin: dict = Depends(auth.require_admin)):
    fs.delete_record("family_fields", field_id)
    return {"ok": True}


# ----------------------- Families -----------------------
@api.get("/families")
def get_families(user: dict = Depends(auth.get_current_user)):
    fams = fs.list_records("families")
    return sorted(fams, key=lambda f: f.get("created_at", ""), reverse=True)


@api.post("/families")
def add_family(body: FamilyInput, user: dict = Depends(auth.get_current_user)):
    rec = fs.push("families", {
        "data": body.data,
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "created_by": user.get("name"),
    })
    return rec


@api.put("/families/{family_id}")
def update_family(family_id: str, body: FamilyInput, user: dict = Depends(auth.get_current_user)):
    if not fs.get_record("families", family_id):
        raise HTTPException(status_code=404, detail="العائلة غير موجودة")
    fs.update_record("families", family_id, {"data": body.data, "updated_at": now_iso()})
    return fs.get_record("families", family_id)


@api.delete("/families/all")
def delete_all_families(admin: dict = Depends(auth.require_admin)):
    families = fs.list_records("families")
    count = len(families)
    # Delete entire collections at once (Firebase node delete)
    fs.ref("families").delete()
    fs.ref("aid_records").delete()
    return {"ok": True, "deleted": count}


@api.delete("/families/{family_id}")
def delete_family(family_id: str, user: dict = Depends(auth.get_current_user)):
    fs.delete_record("families", family_id)
    for r in fs.list_records("aid_records"):
        if r.get("family_id") == family_id:
            fs.delete_record("aid_records", r["id"])
    return {"ok": True}


# ----------------------- Aid Types -----------------------
@api.get("/aid-types")
def get_aid_types(user: dict = Depends(auth.get_current_user)):
    return fs.list_records("aid_types")


@api.post("/aid-types")
def add_aid_type(body: AidTypeCreate, admin: dict = Depends(auth.require_admin)):
    rec = fs.push("aid_types", {
        "name": body.name,
        "description": body.description,
        "created_at": now_iso(),
    })
    return rec


@api.delete("/aid-types/{type_id}")
def delete_aid_type(type_id: str, admin: dict = Depends(auth.require_admin)):
    fs.delete_record("aid_types", type_id)
    return {"ok": True}


# ----------------------- Aid Records -----------------------
@api.get("/aid-records")
def get_aid_records(family_id: Optional[str] = None, user: dict = Depends(auth.get_current_user)):
    records = fs.list_records("aid_records")
    if family_id:
        records = [r for r in records if r.get("family_id") == family_id]
    return sorted(records, key=lambda r: r.get("date", ""), reverse=True)


@api.post("/aid-records")
def add_aid_record(body: AidRecordInput, user: dict = Depends(auth.get_current_user)):
    aid_type = fs.get_record("aid_types", body.aid_type_id)
    rec = fs.push("aid_records", {
        "family_id": body.family_id,
        "aid_type_id": body.aid_type_id,
        "aid_type_name": aid_type.get("name") if aid_type else "",
        "date": body.date,
        "quantity": body.quantity,
        "notes": body.notes,
        "created_by": user.get("name"),
        "created_at": now_iso(),
    })
    return rec


@api.delete("/aid-records/all")
def delete_all_aid_records(admin: dict = Depends(auth.require_admin)):
    records = fs.list_records("aid_records")
    count = len(records)
    fs.ref("aid_records").delete()
    return {"ok": True, "deleted": count}


@api.delete("/aid-records/{record_id}")
def delete_aid_record(record_id: str, user: dict = Depends(auth.get_current_user)):
    fs.delete_record("aid_records", record_id)
    return {"ok": True}


# ----------------------- Stats -----------------------
@api.get("/stats")
def stats(user: dict = Depends(auth.get_current_user)):
    families = fs.list_records("families")
    individual_members = fs.list_records("individual_members")
    records = fs.list_records("aid_records")
    aid_types = fs.list_records("aid_types")
    by_type: Dict[str, int] = {}
    for r in records:
        name = r.get("aid_type_name") or "غير محدد"
        by_type[name] = by_type.get(name, 0) + 1
    recent = sorted(records, key=lambda r: r.get("created_at", ""), reverse=True)[:5]
    return {
        "total_families": len(families),
        "total_individual_members": len(individual_members),
        "total_aid_records": len(records),
        "total_aid_types": len(aid_types),
        "aid_by_type": [{"name": k, "count": v} for k, v in by_type.items()],
        "recent_records": recent,
    }


# ----------------------- Excel Import / Export -----------------------
@api.get("/families/export")
def export_families(user: dict = Depends(auth.get_current_user)):
    fields = sorted(fs.list_records("family_fields"), key=lambda f: f.get("order", 0))
    families = fs.list_records("families")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Families"
    headers = [f["label"] for f in fields]
    ws.append(headers)
    for fam in families:
        data = fam.get("data", {})
        ws.append([data.get(f["key"], "") for f in fields])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=families.xlsx"},
    )


@api.get("/families/template")
def families_template(user: dict = Depends(auth.get_current_user)):
    fields = sorted(fs.list_records("family_fields"), key=lambda f: f.get("order", 0))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Families"
    ws.append([f["label"] for f in fields])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=families_template.xlsx"},
    )


@api.post("/families/import/columns")
async def import_columns(file: UploadFile = File(...), user: dict = Depends(auth.get_current_user)):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="ملف غير صالح. الرجاء رفع ملف Excel (.xlsx)")
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return {"preview": [], "suggested_header": 0, "total_rows": 0}

    preview = []
    for r in all_rows[:15]:
        preview.append(["" if c is None else str(c).strip() for c in r])
    width = max((len(r) for r in preview), default=0)
    preview = [r + [""] * (width - len(r)) for r in preview]

    suggested, best = 0, -1
    for i, r in enumerate(preview):
        cnt = sum(1 for c in r if c)
        if cnt > best:
            best, suggested = cnt, i

    return {"preview": preview, "suggested_header": suggested, "total_rows": len(all_rows)}


@api.post("/families/import")
async def import_families(
    file: UploadFile = File(...),
    mapping: Optional[str] = Form(None),
    header_row: int = Form(0),
    user: dict = Depends(auth.get_current_user),
):
    fields = sorted(fs.list_records("family_fields"), key=lambda f: f.get("order", 0))
    if not fields:
        raise HTTPException(status_code=400, detail="الرجاء إضافة حقول العائلة أولاً قبل الاستيراد")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="ملف غير صالح. الرجاء رفع ملف Excel (.xlsx)")
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows or header_row >= len(all_rows):
        return {"imported": 0}

    header = ["" if c is None else str(c).strip() for c in all_rows[header_row]]

    field_col: Dict[str, int] = {}
    if mapping:
        try:
            mp = json.loads(mapping)
        except Exception:
            mp = {}
        for fkey, col_idx in mp.items():
            if col_idx is None or col_idx == "":
                continue
            try:
                idx = int(col_idx)
            except (TypeError, ValueError):
                continue
            field_col[fkey] = idx
    else:
        label_to_key = {f["label"]: f["key"] for f in fields}
        for idx, col in enumerate(header):
            if col in label_to_key:
                field_col[label_to_key[col]] = idx

    if not field_col:
        raise HTTPException(status_code=400, detail="لم يتم ربط أي عمود بالحقول. الرجاء تحديد الأعمدة المطلوبة")

    imported = 0
    for row in all_rows[header_row + 1:]:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        data = {}
        for fkey, idx in field_col.items():
            val = row[idx] if idx < len(row) else None
            data[fkey] = "" if val is None else str(val).strip()
        if any(str(v).strip() for v in data.values()):
            fs.push("families", {"data": data, "created_at": now_iso(),
                                 "updated_at": now_iso(), "created_by": user.get("name")})
            imported += 1
    return {"imported": imported}


@api.get("/aid-records/export")
def export_aid_records(user: dict = Depends(auth.get_current_user)):
    fields = sorted(fs.list_records("family_fields"), key=lambda f: f.get("order", 0))
    name_field = fields[0]["key"] if fields else None
    families = {f["id"]: f.get("data", {}) for f in fs.list_records("families")}
    records = fs.list_records("aid_records")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AidRecords"
    ws.append(["العائلة", "نوع المساعدة", "التاريخ", "الكمية", "ملاحظات", "أُضيف بواسطة"])
    for r in records:
        fam_data = families.get(r.get("family_id"), {})
        fam_name = fam_data.get(name_field, r.get("family_id")) if name_field else r.get("family_id")
        ws.append([fam_name, r.get("aid_type_name"), r.get("date"), r.get("quantity"),
                   r.get("notes"), r.get("created_by")])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=aid_records.xlsx"},
    )


@api.post("/aid-records/import")
async def import_aid_records(
    file: UploadFile = File(...),
    header_row: int = Form(0),
    match_column: int = Form(...),
    match_field_key: str = Form(...),
    aid_type_id: str = Form(...),
    date: str = Form(...),
    notes: Optional[str] = Form(""),
    fuzzy: bool = Form(True),          # enable fuzzy Arabic matching
    threshold: float = Form(0.60),      # minimum match score (0-1)
    user: dict = Depends(auth.get_current_user),
):
    aid_type = fs.get_record("aid_types", aid_type_id)
    if not aid_type:
        raise HTTPException(status_code=400, detail="نوع المساعدة غير موجود")

    families = fs.list_records("families")

    # Build exact-match index (normalized) for fast lookup
    exact_index: Dict[str, str] = {}
    for fam in families:
        val = str(fam.get("data", {}).get(match_field_key, "")).strip()
        if val:
            norm = normalize_arabic(val)
            if norm and norm not in exact_index:
                exact_index[norm] = fam["id"]

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="ملف غير صالح. الرجاء رفع ملف Excel (.xlsx)")
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    created = 0
    unmatched: List[str] = []
    fuzzy_matched = 0

    for row in all_rows[header_row + 1:]:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        ident = row[match_column] if match_column < len(row) else None
        ident = "" if ident is None else str(ident).strip()
        if not ident:
            continue

        # 1. Exact normalized match
        fam_id = exact_index.get(normalize_arabic(ident))

        # 2. Fuzzy match if enabled and no exact match found
        if not fam_id and fuzzy:
            best_fam, score = find_best_family(families, match_field_key, ident, threshold)
            if best_fam:
                fam_id = best_fam["id"]
                fuzzy_matched += 1

        if not fam_id:
            unmatched.append(ident)
            continue

        fs.push("aid_records", {
            "family_id": fam_id,
            "aid_type_id": aid_type_id,
            "aid_type_name": aid_type.get("name"),
            "date": date,
            "quantity": "",
            "notes": notes,
            "created_by": user.get("name"),
            "created_at": now_iso(),
        })
        created += 1

    return {
        "created": created,
        "fuzzy_matched": fuzzy_matched,
        "unmatched_count": len(unmatched),
        "unmatched": unmatched[:50],
    }


# ----------------------- Categories (special groups) -----------------------
DEFAULT_CATEGORIES = [
    {"name": "أرامل", "key": "widows", "icon": "HeartHandshake", "order": 1},
    {"name": "حوامل", "key": "pregnant", "icon": "Baby", "order": 2},
    {"name": "مرضعات", "key": "nursing", "icon": "Milk", "order": 3},
    {"name": "أطفال", "key": "children", "icon": "ToyBrick", "order": 4},
    {"name": "مرضى", "key": "patients", "icon": "Stethoscope", "order": 5},
    {"name": "كبار السن", "key": "elderly", "icon": "Accessibility", "order": 6},
    {"name": "إصابات", "key": "injuries", "icon": "Bandage", "order": 7},
]


def seed_categories():
    existing = fs.list_records("categories")
    existing_keys = {c.get("key") for c in existing}
    for c in DEFAULT_CATEGORIES:
        if c["key"] not in existing_keys:
            fs.push("categories", {**c, "system": True, "created_at": now_iso()})


class CategoryCreate(BaseModel):
    name: str
    icon: Optional[str] = "Layers"


class CategoryFieldCreate(BaseModel):
    category_id: str
    label: str
    key: Optional[str] = None
    type: str = "text"
    order: int = 0


class CategoryRecordInput(BaseModel):
    category_id: str
    family_id: Optional[str] = ""
    data: Dict[str, Any] = {}


@api.get("/categories")
def get_categories(user: dict = Depends(auth.get_current_user)):
    cats = fs.list_records("categories")
    records = fs.list_records("category_records")
    counts: Dict[str, int] = {}
    for r in records:
        cid = r.get("category_id")
        counts[cid] = counts.get(cid, 0) + 1
    for c in cats:
        c["count"] = counts.get(c["id"], 0)
    return sorted(cats, key=lambda c: c.get("order", 999))


@api.post("/categories")
def add_category(body: CategoryCreate, admin: dict = Depends(auth.require_admin)):
    existing = fs.list_records("categories")
    order = max([c.get("order", 0) for c in existing], default=0) + 1
    rec = fs.push("categories", {
        "name": body.name,
        "key": f"cat_{int(datetime.now().timestamp()*1000)}",
        "icon": body.icon or "Layers",
        "order": order,
        "system": False,
        "created_at": now_iso(),
    })
    return rec


@api.delete("/categories/{category_id}")
def delete_category(category_id: str, admin: dict = Depends(auth.require_admin)):
    fs.delete_record("categories", category_id)
    for f in fs.list_records("category_fields"):
        if f.get("category_id") == category_id:
            fs.delete_record("category_fields", f["id"])
    for r in fs.list_records("category_records"):
        if r.get("category_id") == category_id:
            fs.delete_record("category_records", r["id"])
    return {"ok": True}


@api.get("/category-fields")
def get_category_fields(category_id: str, user: dict = Depends(auth.get_current_user)):
    fields = [f for f in fs.list_records("category_fields") if f.get("category_id") == category_id]
    return sorted(fields, key=lambda f: f.get("order", 0))


@api.post("/category-fields")
def add_category_field(body: CategoryFieldCreate, admin: dict = Depends(auth.require_admin)):
    key = body.key or f"cf_{int(datetime.now().timestamp()*1000)}"
    rec = fs.push("category_fields", {
        "category_id": body.category_id,
        "label": body.label,
        "key": key,
        "type": body.type,
        "order": body.order,
        "created_at": now_iso(),
    })
    return rec


@api.delete("/category-fields/{field_id}")
def delete_category_field(field_id: str, admin: dict = Depends(auth.require_admin)):
    fs.delete_record("category_fields", field_id)
    return {"ok": True}


@api.get("/category-records")
def get_category_records(category_id: str, user: dict = Depends(auth.get_current_user)):
    records = [r for r in fs.list_records("category_records") if r.get("category_id") == category_id]
    return sorted(records, key=lambda r: r.get("created_at", ""), reverse=True)


@api.post("/category-records")
def add_category_record(body: CategoryRecordInput, user: dict = Depends(auth.get_current_user)):
    rec = fs.push("category_records", {
        "category_id": body.category_id,
        "family_id": body.family_id,
        "data": body.data,
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "created_by": user.get("name"),
    })
    return rec


@api.put("/category-records/{record_id}")
def update_category_record(record_id: str, body: CategoryRecordInput, user: dict = Depends(auth.get_current_user)):
    if not fs.get_record("category_records", record_id):
        raise HTTPException(status_code=404, detail="السجل غير موجود")
    fs.update_record("category_records", record_id, {
        "family_id": body.family_id,
        "data": body.data,
        "updated_at": now_iso(),
    })
    return fs.get_record("category_records", record_id)


@api.delete("/category-records/all")
def delete_all_category_records(category_id: str, admin: dict = Depends(auth.require_admin)):
    records = [r for r in fs.list_records("category_records") if r.get("category_id") == category_id]
    for r in records:
        fs.delete_record("category_records", r["id"])
    return {"ok": True, "deleted": len(records)}


@api.delete("/category-records/{record_id}")
def delete_category_record(record_id: str, user: dict = Depends(auth.get_current_user)):
    fs.delete_record("category_records", record_id)
    return {"ok": True}


@api.get("/category-records/export")
def export_category_records(category_id: str, user: dict = Depends(auth.get_current_user)):
    category = fs.get_record("categories", category_id)
    cat_fields = sorted(
        [f for f in fs.list_records("category_fields") if f.get("category_id") == category_id],
        key=lambda f: f.get("order", 0),
    )
    fam_fields = sorted(fs.list_records("family_fields"), key=lambda f: f.get("order", 0))
    name_field = fam_fields[0]["key"] if fam_fields else None
    families = {f["id"]: f.get("data", {}) for f in fs.list_records("families")}
    records = [r for r in fs.list_records("category_records") if r.get("category_id") == category_id]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Records"
    ws.append(["الاسم (العائلة)"] + [f["label"] for f in cat_fields])
    for r in records:
        fam_data = families.get(r.get("family_id"), {})
        fam_name = fam_data.get(name_field, "") if name_field else ""
        ws.append([fam_name] + [r.get("data", {}).get(f["key"], "") for f in cat_fields])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = (category.get("name") if category else "category") + ".xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=category.xlsx"},
    )


@api.post("/category-records/import")
async def import_category_records(
    file: UploadFile = File(...),
    category_id: str = Form(...),
    header_row: int = Form(0),
    match_column: int = Form(...),
    match_field_key: str = Form(...),
    mapping: Optional[str] = Form(None),
    fuzzy: bool = Form(True),
    threshold: float = Form(0.60),
    user: dict = Depends(auth.get_current_user),
):
    families = fs.list_records("families")
    exact_index: Dict[str, str] = {}
    for fam in families:
        val = str(fam.get("data", {}).get(match_field_key, "")).strip()
        if val:
            norm = normalize_arabic(val)
            if norm and norm not in exact_index:
                exact_index[norm] = fam["id"]

    field_col: Dict[str, int] = {}
    if mapping:
        try:
            mp = json.loads(mapping)
        except Exception:
            mp = {}
        for fkey, col_idx in mp.items():
            if col_idx is None or col_idx == "":
                continue
            try:
                field_col[fkey] = int(col_idx)
            except (TypeError, ValueError):
                continue

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="ملف غير صالح. الرجاء رفع ملف Excel (.xlsx)")
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    created = 0
    fuzzy_matched = 0
    unmatched: List[str] = []

    for row in all_rows[header_row + 1:]:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        ident = row[match_column] if match_column < len(row) else None
        ident = "" if ident is None else str(ident).strip()
        if not ident:
            continue

        fam_id = exact_index.get(normalize_arabic(ident))
        if not fam_id and fuzzy:
            best_fam, score = find_best_family(families, match_field_key, ident, threshold)
            if best_fam:
                fam_id = best_fam["id"]
                fuzzy_matched += 1
        if not fam_id:
            unmatched.append(ident)
            continue

        data = {}
        for fkey, idx in field_col.items():
            val = row[idx] if idx < len(row) else None
            data[fkey] = "" if val is None else str(val).strip()

        fs.push("category_records", {
            "category_id": category_id,
            "family_id": fam_id,
            "data": data,
            "created_at": now_iso(),
            "updated_at": now_iso(),
            "created_by": user.get("name"),
        })
        created += 1

    return {
        "created": created,
        "fuzzy_matched": fuzzy_matched,
        "unmatched_count": len(unmatched),
        "unmatched": unmatched[:50],
    }


@api.get("/")
def root():
    return {"message": "مخيم العائدين API"}


app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def on_startup():
    auth.seed_admin()
    seed_categories()
    logger.info("Admin seeded. Firebase ready.")
